from __future__ import annotations

from typing import Dict, List, Tuple

from openpyxl import load_workbook

from src.services.rules.exceptions import RulesWorkbookError
from src.services.rules.rules_types import (
    RulesWorkbook,
    RulesProfile,
    MappingRow,
    DefaultChildRow,
)


def _norm(v) -> str:
    return (str(v).strip() if v is not None else "").strip()


def _to_int(v, default: int = 0) -> int:
    try:
        return int(str(v).strip())
    except Exception:
        return default


def _read_table(ws) -> Tuple[List[str], List[List[str]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    def norm_row(r):
        return [(_norm(x)) for x in (r or [])]

    hdr = norm_row(rows[0])
    data = [norm_row(r) for r in rows[1:]]

    def rtrim(a: List[str]) -> List[str]:
        b = list(a)
        while b and (b[-1] or "").strip() == "":
            b.pop()
        return b

    hdr = rtrim(hdr)
    data = [rtrim(r) for r in data if any((c or "").strip() for c in r)]
    return hdr, data


def _idx_map(headers: List[str]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for i, h in enumerate(headers or []):
        key = (h or "").strip().lower()
        if key and key not in out:
            out[key] = i
    return out


def _get(row: List[str], idx: Dict[str, int], name: str, default: str = "") -> str:
    j = idx.get((name or "").strip().lower(), None)
    if j is None:
        return default
    if j >= len(row):
        return default
    return row[j]


def _parse_profiles_format(wb) -> RulesWorkbook:
    """
    Legacy 'rules.xlsx' format:
      - Profiles
      - (optional) LevelMappings
      - (optional) DefaultChildren
    """
    out = RulesWorkbook()

    if "Profiles" not in wb.sheetnames:
        raise RulesWorkbookError(
            message="Invalid rules workbook format (missing required sheet).",
            expected_sheets=["Profiles"],
            found_sheets=list(wb.sheetnames),
        )

    ws = wb["Profiles"]
    headers, data = _read_table(ws)
    idx = _idx_map(headers)

    for r in data:
        pid = _norm(_get(r, idx, "profile_id"))
        if not pid:
            continue

        pname = _norm(_get(r, idx, "profile_name")) or pid
        level_count = _to_int(_get(r, idx, "level_count"), 2)

        level_labels: Dict[int, str] = {}
        for lv in range(1, max(1, level_count) + 1):
            level_labels[lv] = _norm(_get(r, idx, f"level_{lv}_label")) or f"Level {lv}"

        delim = _norm(_get(r, idx, "code_delimiter")) or "."
        if len(delim) != 1:
            delim = "."

        prof = RulesProfile(
            profile_id=pid,
            profile_name=pname,
            level_count=level_count,
            level_labels=level_labels,
            code_delimiter=delim,
            code_regex=_norm(_get(r, idx, "code_regex")),
            notes=_norm(_get(r, idx, "notes")),
        )
        out.profiles[pid] = prof

    if "LevelMappings" in wb.sheetnames:
        ws = wb["LevelMappings"]
        headers, data = _read_table(ws)
        idx = _idx_map(headers)

        for r in data:
            pid = _norm(_get(r, idx, "profile_id"))
            if not pid:
                continue
            level = _to_int(_get(r, idx, "level"), 0)
            code = _norm(_get(r, idx, "code"))
            name = _norm(_get(r, idx, "name"))
            locked = (_norm(_get(r, idx, "locked")).lower() in ("1", "true", "yes", "y"))

            if level <= 0 or not code:
                continue

            mr = MappingRow(profile_id=pid, level=level, code=code, name=name, locked=locked)
            out.mappings[(pid, level, code.strip().lower())] = mr

    if "DefaultChildren" in wb.sheetnames:
        ws = wb["DefaultChildren"]
        headers, data = _read_table(ws)
        idx = _idx_map(headers)

        for r in data:
            pid = _norm(_get(r, idx, "profile_id"))
            if not pid:
                continue
            parent_level = _to_int(_get(r, idx, "parent_level"), 0)
            parent_code = _norm(_get(r, idx, "parent_code"))
            child_level = _to_int(_get(r, idx, "child_level"), 0)
            child_code = _norm(_get(r, idx, "child_code"))
            child_name = _norm(_get(r, idx, "child_name"))

            if parent_level <= 0 or child_level <= 0 or not parent_code or not child_code:
                continue

            dc = DefaultChildRow(
                profile_id=pid,
                parent_level=parent_level,
                parent_code=parent_code,
                child_level=child_level,
                child_code=child_code,
                child_name=child_name,
            )
            out.default_children.append(dc)

    return out


def _parse_simple_authority_format(wb) -> RulesWorkbook:
    """
    'Human' workbook:
      - Profile sheet: Field | Value
      - Level 1 sheet: Code | Name
      - Level 2 sheet: Parent code | Code (suffix) | Name
      - Optional Level 3..N
    """
    out = RulesWorkbook()

    if ("Profile" not in wb.sheetnames) or ("Level 1" not in wb.sheetnames):
        raise RulesWorkbookError(
            message="Invalid simple authority workbook format.",
            expected_sheets=["Profile", "Level 1"],
            found_sheets=list(wb.sheetnames),
        )

    ws = wb["Profile"]
    headers, data = _read_table(ws)

    kv: Dict[str, str] = {}
    for r in ([headers] + data):
        if len(r) < 2:
            continue
        k = _norm(r[0]).lower()
        v = _norm(r[1])
        if k:
            kv[k] = v

    discipline = (kv.get("discipline", "") or "").strip()
    delim = (kv.get("code delimiter", ".") or ".").strip() or "."
    if len(delim) != 1:
        delim = "."

    level_count = _to_int(kv.get("number of levels", 2), 2)
    level_count = max(2, min(level_count, 6))

    pad_l1 = _to_int(kv.get("pad level 1 codes to", 2), 2)
    pad_l1 = max(1, min(pad_l1, 6))

    level_labels: Dict[int, str] = {}
    for lv in range(1, level_count + 1):
        nm = kv.get(f"level {lv} name", "") or kv.get(f"level {lv} label", "")
        level_labels[lv] = (nm or f"Level {lv}").strip()

    pid = "default"
    pname = (discipline or "Authority Profile").strip() or "Authority Profile"
    out.profiles[pid] = RulesProfile(
        profile_id=pid,
        profile_name=pname,
        level_count=level_count,
        level_labels=level_labels,
        code_delimiter=delim,
        code_regex="",
        notes="",
    )

    def norm_l1(c: str) -> str:
        c = (c or "").strip()
        if c.isdigit():
            return str(int(c)).zfill(pad_l1)
        return c

    ws = wb["Level 1"]
    headers, data = _read_table(ws)
    idx = _idx_map(headers)
    use_headers = ("code" in idx and "name" in idx)

    for r in data:
        if use_headers:
            code = _norm(_get(r, idx, "code"))
            name = _norm(_get(r, idx, "name"))
        else:
            code = _norm(r[0] if len(r) > 0 else "")
            name = _norm(r[1] if len(r) > 1 else "")

        if not code:
            continue
        code = norm_l1(code)

        mr = MappingRow(profile_id=pid, level=1, code=code, name=name, locked=True)
        out.mappings[(pid, 1, code.lower())] = mr

    for lv in range(2, level_count + 1):
        sheet = f"Level {lv}"
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]
        headers, data = _read_table(ws)
        idx = _idx_map(headers)
        use_headers = ("parent code" in idx and "code" in idx and "name" in idx)

        for r in data:
            if use_headers:
                parent = _norm(_get(r, idx, "parent code"))
                suffix = _norm(_get(r, idx, "code"))
                name = _norm(_get(r, idx, "name"))
            else:
                parent = _norm(r[0] if len(r) > 0 else "")
                suffix = _norm(r[1] if len(r) > 1 else "")
                name = _norm(r[2] if len(r) > 2 else "")

            if not parent or not suffix:
                continue

            parent = norm_l1(parent) if lv == 2 else parent.strip()
            full = suffix if (delim in suffix) else f"{parent}{delim}{str(int(suffix)) if suffix.isdigit() else suffix}"

            mr = MappingRow(profile_id=pid, level=lv, code=full, name=name, locked=True)
            out.mappings[(pid, lv, full.lower())] = mr

    return out


def load_rules_workbook(path: str) -> RulesWorkbook:
    wb = load_workbook(path, data_only=True)
    if "Profiles" in wb.sheetnames:
        return _parse_profiles_format(wb)
    return _parse_simple_authority_format(wb)


# Backwards-compatible API: UI imports this name.
def load_rules_xlsx(path: str) -> RulesWorkbook:
    return load_rules_workbook(path)
